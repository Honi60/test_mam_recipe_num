import os
import json
import tkinter as tk
from tkinter import ttk, messagebox

DB_DIR = r"G:\My Drive\Rentals\RentalsDB"
HISTORY_FILE = os.path.join(DB_DIR, "history.json")


class ToExcelApp(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        self.pack(fill="both", expand=True, padx=10, pady=10)
        os.makedirs(DB_DIR, exist_ok=True)
        self.history = {}
        self.load_history()
        self.build_ui()

    def load_history(self):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                self.history = json.load(f)
        except Exception:
            self.history = {}

    def reload_history(self):
        """Reload history from disk and write a short message to the UI log."""
        # Keep reload silent for UI - update internal cache only
        self.load_history()

    def build_ui(self):
        row = tk.Frame(self)
        row.pack(fill="x", pady=6)
        tk.Label(row, text="Month (1-12):").pack(side="left")
        self.month_var = tk.StringVar()
        tk.Entry(row, textvariable=self.month_var, width=6).pack(side="left", padx=(4,8))
        tk.Label(row, text="Year (e.g. 2025):").pack(side="left")
        self.year_var = tk.StringVar()
        tk.Entry(row, textvariable=self.year_var, width=8).pack(side="left", padx=(4,8))
        tk.Button(row, text="Export to Clipboard", command=self.export).pack(side="left")
        tk.Button(row, text="Export .xlsx", command=self.export_xlsx).pack(side="left", padx=(8,0))

        self.log = tk.Text(self, height=20)
        self.log.pack(fill="both", expand=True, pady=(8,0))

    def _collect_sorted_rows(self, month, year):
        """Yield (date_obj, row_values) for entries matching month/year, sorted by date ascending."""
        items = []
        from datetime import date
        for key in sorted(self.history.keys()):
            entry = self.history.get(key, {})
            if not entry or not isinstance(entry, dict):
                continue
            cust = list(entry.keys())[0]
            data = entry[cust]
            date_str = data.get("Date", "")
            try:
                parts = date_str.split('/')
                if len(parts) == 3:
                    d = int(parts[0])
                    m = int(parts[1])
                    y = int(parts[2])
                    if y < 100:
                        y += 2000
                else:
                    continue
            except Exception:
                continue
            if m != month or y != year:
                continue
            try:
                dt = date(y, m, d)
            except Exception:
                continue
            payment = data.get("payment", "")
            bank_number = data.get("BankNumber", "")
            bank_account = data.get("bankAccount", data.get("bankAccount", ""))
            receipt_no = data.get("recipeNum", key)
            # Columns: A(hebrew 'הכנסה'), B(date), C(payment), D(customer), E(blank), F(receipt_no), G(bank_number), H(bank_account), I(CheckNumber)
            check_number = data.get("CheckNumber", "")
            row_vals = ["הכנסה", date_str, payment, cust, "", receipt_no, bank_number, bank_account, check_number]
            items.append((dt, row_vals))
        items.sort(key=lambda x: x[0])
        for it in items:
            yield it

    def export(self):
        month_s = self.month_var.get().strip()
        year_s = self.year_var.get().strip()
        if not month_s or not year_s:
            messagebox.showwarning("Input required", "Please enter month and year.")
            return
        try:
            month = int(month_s)
            year = int(year_s)
        except Exception:
            messagebox.showerror("Error", "Month and year must be integers.")
            return
        rows = [r for _, r in self._collect_sorted_rows(month, year)]
        if not rows:
            messagebox.showinfo("No data", "No receipts found for that month/year.")
            return
        # build TSV
        tsv_lines = ["\t".join(map(str, r)) for r in rows]
        tsv = "\n".join(tsv_lines)
        # copy to clipboard
        try:
            self.master.clipboard_clear()
            self.master.clipboard_append(tsv)
            self.log.delete(1.0, tk.END)
            self.log.insert(tk.END, f"Copied {len(rows)} rows to clipboard.\n")
            self.log.insert(tk.END, tsv)
            messagebox.showinfo("Copied", f"Copied {len(rows)} rows to clipboard.")
        except Exception as e:
            messagebox.showerror("Clipboard Error", f"Failed to copy to clipboard: {e}")

    def export_xlsx(self):
        # collect the same rows as export()
        month_s = self.month_var.get().strip()
        year_s = self.year_var.get().strip()
        if not month_s or not year_s:
            messagebox.showwarning("Input required", "Please enter month and year.")
            return
        try:
            month = int(month_s)
            year = int(year_s)
        except Exception:
            messagebox.showerror("Error", "Month and year must be integers.")
            return
        rows = [r for _, r in self._collect_sorted_rows(month, year)]
        if not rows:
            messagebox.showinfo("No data", "No receipts found for that month/year.")
            return
        # ask where to save
        fpath = tk.filedialog.asksaveasfilename(title="Save Excel File", defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if not fpath:
            return
        # ensure openpyxl available
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showerror("Missing dependency", "openpyxl is required to export XLSX. Install with: pip install openpyxl")
            return
        # write workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        for r_idx, r in enumerate(rows, start=1):
            for c_idx, val in enumerate(r, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        try:
            wb.save(fpath)
            messagebox.showinfo("Saved", f"Saved {len(rows)} rows to {fpath}")
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save Excel file: {e}")


if __name__ == '__main__':
    root = tk.Tk()
    root.title("Export History to Excel Clipboard")
    app = ToExcelApp(root)
    root.geometry('800x600')
    root.mainloop()
